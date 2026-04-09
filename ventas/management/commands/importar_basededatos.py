from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal, InvalidOperation
from catalogos.models import Categoria, Sucursal, MetodoPago, Cliente, Vendedor, Producto
from ventas.models import Venta, DetalleVenta

class Command(BaseCommand):
    help = 'Importa datos desde DashBoard2021.xlsx (versión robusta)'

    def add_arguments(self, parser):
        parser.add_argument('--archivo', type=str, required=True)

    def handle(self, *args, **options):
        archivo = options['archivo']
        self.stdout.write(f" Cargando: {archivo}")
        
        wb = load_workbook(archivo, data_only=True)
        
        # Buscar hoja correcta (ignora mayúsculas/minúsculas)
        sheet_name = None
        for name in wb.sheetnames:
            if 'basededatos' in name.lower():
                sheet_name = name
                break
        if not sheet_name:
            self.stdout.write(self.style.ERROR("❌ No se encontró la hoja 'BaseDeDatos'"))
            return

        ws = wb[sheet_name]
        filas = list(ws.iter_rows(values_only=True))

        # 🔍 1. DETECCIÓN DINÁMICA DE ENCABEZADOS
        header_idx = 0
        for i, fila in enumerate(filas[:10]):  # Revisa las primeras 10 filas
            if fila and any(c and 'fecha' in str(c).lower() for c in fila):
                header_idx = i
                self.stdout.write(f" Encabezados encontrados en fila Excel #{i+1}")
                break

        encabezados_raw = filas[header_idx]
        # Crear diccionario: nombre_normalizado -> índice_de_columna
        col_map = {}
        for idx, col_name in enumerate(encabezados_raw):
            if col_name:
                clean = str(col_name).strip().lower()
                col_map[clean] = idx

        self.stdout.write(f" Columnas detectadas: {list(col_map.keys())}")

        datos = filas[header_idx+1:]
        ventas_temporales = {}
        filas_procesadas = 0
        filas_omitidas = 0

        # Funciones auxiliares seguras
        def get_val(key):
            idx = col_map.get(key.lower())
            if idx is None or idx >= len(fila) or fila[idx] is None:
                return None
            return fila[idx]

        def get_str(key):
            val = get_val(key)
            return str(val).strip() if val is not None else ''

        def get_decimal(key):
            val = get_val(key)
            if val is None or str(val).strip() == '':
                return Decimal('0')
            limpio = str(val).replace(',', '').replace(' ', '')
            try:
                return Decimal(limpio)
            except InvalidOperation:
                return Decimal('0')

        for idx_fila, fila in enumerate(datos, start=header_idx+2):
            if not fila or not any(fila):
                continue

            # --- EXTRACCIÓN ROBUSTA ---
            folio = get_str('documento')
            fecha_valor = get_val('fecha')
            cliente_nombre = get_str('cliente')
            ciudad = get_str('ciudad')
            estado = get_str('provincia')
            pais = 'Ecuador'  # Ajustar si es otro país
            vendedor_nombre = get_str('vendedor')
            metodo_pago_nombre = get_str('forma de pago')
            nombre_producto = get_str('producto')
            categoria_nombre = get_str('categoría')
            cantidad = get_decimal('cantidad')
            precio_unitario = get_decimal('precio')
            importe = get_decimal('ventas')

            # --- PARSEO SEGURO DE FECHA ---
            fecha = None
            if isinstance(fecha_valor, datetime):
                fecha = fecha_valor.date()
            elif fecha_valor:
                fecha_str = str(fecha_valor).strip()
                try:
                    fecha = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                except ValueError:
                    try:
                        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                    except ValueError:
                        self.stdout.write(self.style.WARNING(f" Fila {idx_fila}: Fecha inválida '{fecha_str}'"))
                        filas_omitidas += 1
                        continue
            else:
                self.stdout.write(self.style.WARNING(f" Fila {idx_fila}: Fecha vacía"))
                filas_omitidas += 1
                continue

            # Cálculo de impuestos (16%)
            subtotal = importe / Decimal('1.16') if importe > 0 else Decimal('0')
            impuesto = importe - subtotal
            total = importe

            # --- CATÁLOGOS ---
            categoria, _ = Categoria.objects.get_or_create(nombre=categoria_nombre or 'General')
            sucursal, _ = Sucursal.objects.get_or_create(
                nombre=cliente_nombre,
                defaults={'ciudad': ciudad, 'estado': estado, 'pais': pais}
            )
            metodo_pago, _ = MetodoPago.objects.get_or_create(nombre=metodo_pago_nombre or 'Efectivo')
            cliente, _ = Cliente.objects.get_or_create(nombre=cliente_nombre)
            vendedor, _ = Vendedor.objects.get_or_create(nombre=vendedor_nombre, defaults={'sucursal': sucursal})
            producto, _ = Producto.objects.get_or_create(
                codigo=f"PROD-{nombre_producto[:15].replace(' ', '')}",
                defaults={'nombre': nombre_producto, 'categoria': categoria, 'precio': precio_unitario}
            )

            # --- AGRUPAR POR FOLIO ---
            if folio not in ventas_temporales:
                ventas_temporales[folio] = {
                    'fecha': fecha, 'cliente': cliente, 'sucursal': sucursal,
                    'vendedor': vendedor, 'metodo_pago': metodo_pago,
                    'subtotal': subtotal, 'impuesto': impuesto, 'total': total,
                    'detalles': []
                }
            ventas_temporales[folio]['detalles'].append({
                'producto': producto, 'cantidad': cantidad,
                'precio_unitario': precio_unitario, 'importe': importe
            })
            filas_procesadas += 1

        # --- GUARDAR EN BD ---
        self.stdout.write(" Guardando ventas y detalles en la base de datos...")
        for folio, info in ventas_temporales.items():
            venta, creada = Venta.objects.get_or_create(
                folio=folio,
                defaults={k: v for k, v in info.items() if k != 'detalles'}
            )
            if creada:
                for d in info['detalles']:
                    DetalleVenta.objects.create(venta=venta, **d)

        self.stdout.write(self.style.SUCCESS(f" Importación finalizada."))
        self.stdout.write(self.style.SUCCESS(f" {filas_procesadas} filas procesadas | {filas_omitidas} omitidas."))
        self.stdout.write(self.style.SUCCESS(f" {len(ventas_temporales)} ventas creadas."))